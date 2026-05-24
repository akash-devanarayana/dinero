"""Import historical monthly data from the Excel workbook into the SQLite DB.

The workbook has one sheet per month (October 2024 .. May 2026). Layouts drift
across months (column offsets shift, section header labels get reused), so the
parser classifies each section by the *set* of header labels it contains rather
than trusting a fixed column position.

Run:  python backend/import_excel.py [path/to/workbook.xlsx]
Re-running is safe: it wipes and rebuilds all tables.
"""

import datetime
import os
import re
import sys

import openpyxl

from db import get_conn, init_db

DEFAULT_XLSX = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "2024-2025-2026 _ Monthly Expenses.xlsx",
)

# Reference date used to decide whether an unpaid bill is 'due' (future) or
# 'over' (past). Matches the app's notion of "today".
TODAY = datetime.date(2026, 5, 24)

MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def norm(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip().lower()


def to_iso(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return None


def to_num(v):
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def sheet_period(name):
    """Return 'YYYY-MM' for a monthly sheet name, else None for non-month sheets."""
    low = name.lower()
    # explicit year suffix, e.g. "JAN - 2026" / "January - 2025"
    m = re.search(r"([a-z]{3,9}).*?(\d{4})", low)
    if m and m.group(1)[:3] in MONTHS:
        return f"{int(m.group(2))}-{MONTHS[m.group(1)[:3]]:02d}"
    # bare month name with no year -> the 2024 block (October/November/December)
    key = low.strip()[:3]
    if key in MONTHS and not any(c.isdigit() for c in low):
        return f"2024-{MONTHS[key]:02d}"
    return None


def normalize_status(raw, due_iso, paid_iso):
    s = norm(raw)
    if s in ("paid", "settled"):
        return "paid"
    if s in ("n/a", "na", ""):
        return "na" if not paid_iso else "paid"
    # unpaid: over if its due date is in the past, otherwise still due
    if due_iso:
        try:
            due = datetime.date.fromisoformat(due_iso)
            return "over" if due < TODAY else "due"
        except ValueError:
            pass
    return "over"


# Header-label vocabulary used to classify a row as a section header.
AMOUNT_LABELS = {"outstanding amount", "amount", "billed amount"}
MIN_LABELS = {"minimum amount", "minimum payment"}


def grid(ws, max_cols=16):
    rows = []
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_cols, values_only=True):
        rows.append(list(r))
    return rows


def label_set(row):
    return {norm(c) for c in row if norm(c)}


def first_text_col(row):
    for i, c in enumerate(row):
        if norm(c):
            return i
    return None


def col_of(row, *labels):
    """Index of the first cell in row matching any of the given normalized labels."""
    targets = set(labels)
    for i, c in enumerate(row):
        if norm(c) in targets:
            return i
    return None


METER_NAMES = {"electricity", "water", "gas", "lpg"}


def classify_header(row):
    """Return a section type for a header row, or None if it isn't one."""
    labels = label_set(row)
    has_status = "status" in labels
    has_amount = bool(labels & AMOUNT_LABELS)
    has_outstanding = bool(labels & {"outstanding amount", "billed amount"})
    has_min = bool(labels & MIN_LABELS)
    has_due = "due date" in labels
    is_merchant = bool(labels & {"merchant", "amount to be paid"}) or any(
        "pay by 3" in l for l in labels
    )
    is_sub = any("subscription" in l for l in labels)
    has_units = bool(labels & {"units", "units consumed"})

    if has_units and not has_amount and not has_outstanding:
        return "meter"  # standalone units block (handled by a separate scan)
    if is_merchant:
        return "merchant"
    if has_outstanding and has_min:
        return "card"
    if is_sub:
        return "subscription"
    # loan header: amount + status only (no due date, no outstanding, no minimum).
    # Older months reuse the literal "Card" label for this, so check before utility.
    if {"amount", "status"} <= labels and not has_due and not has_outstanding and not has_min:
        return "loan"
    if has_outstanding or has_amount:
        return "utility"
    return None


def is_blank(row):
    return all(norm(c) == "" for c in row)


def parse_sheet(rows, period):
    out = {"bills": [], "meters": [], "loans": []}
    consumed = [False] * len(rows)

    # ── Meters: locate any 'units'/'units consumed' header anywhere in the sheet.
    for r, row in enumerate(rows):
        uc = col_of(row, "units", "units consumed")
        if uc is None:
            continue
        # name column sits just left of the units header when that cell is a
        # section label (later layouts); otherwise the units header column itself
        # holds the names (October's "Units Consumed").
        left = norm(row[uc - 1]) if uc > 0 else ""
        if left in ("utility", "card", "electricity", "water"):
            name_col, val_col = uc - 1, uc
        else:
            name_col, val_col = uc, uc + 1
        order = 0
        for rr in range(r + 1, len(rows)):
            nm = rows[rr][name_col] if name_col < len(rows[rr]) else None
            val = rows[rr][val_col] if val_col < len(rows[rr]) else None
            nname = norm(nm)
            if nname in ("", "utility", "card"):
                continue
            if nname not in METER_NAMES:
                # next section (e.g. loans sit directly below with no blank row)
                break
            reading = to_num(val)
            if reading is None:
                continue
            unit = "m³" if "water" in nname else "kWh"
            out["meters"].append({
                "name": str(nm).strip(), "unit": unit, "reading": reading,
                "reading_date": None, "note": None, "sort_order": order,
            })
            order += 1

    # ── Sectioned bills + loans.
    section = None
    cols = {}
    order = 0
    for r, row in enumerate(rows):
        kind = classify_header(row)
        if kind:
            section = kind
            order = 0
            name_col = first_text_col(row)
            cols = {
                "name": name_col,
                "amount": col_of(row, "outstanding amount", "amount", "billed amount"),
                "min": col_of(row, "minimum amount", "minimum payment"),
                "due": col_of(row, "due date"),
                "status": col_of(row, "status"),
                "method": col_of(row, "payment method"),
                "paid_on": col_of(row, "date paid"),
                "paid_amount": col_of(row, "paid amount"),
            }
            continue
        if section in (None, "merchant", "meter", "loan"):
            # loans are gathered by the marker-based scan below; this guard just
            # prevents loan rows from being mis-parsed as a utility section.
            continue
        if is_blank(row):
            section = None
            continue

        name_col = cols.get("name")
        if name_col is None:
            continue
        name = row[name_col] if name_col < len(row) else None
        nname = norm(name)
        if nname == "" or nname in ("utility", "card", "subscription", "merchant"):
            continue
        # a "<X> Total" / remainder summary cell is not a record
        if nname.endswith("total") or "remainder" in nname:
            continue

        def cell(key):
            i = cols.get(key)
            return row[i] if i is not None and i < len(row) else None

        amount = to_num(cell("amount")) or 0.0
        due_iso = to_iso(cell("due"))
        paid_iso = to_iso(cell("paid_on"))
        category = {"card": "Card", "utility": "Utility", "subscription": "Subscription"}[section]
        out["bills"].append({
            "category": category,
            "name": str(name).strip(),
            "amount": amount,
            "min_amount": to_num(cell("min")),
            "due_date": due_iso,
            "status": normalize_status(cell("status"), due_iso, paid_iso),
            "payment_method": (str(cell("method")).strip() if cell("method") else None),
            "paid_on": paid_iso,
            "paid_amount": to_num(cell("paid_amount")),
            "cadence": "monthly" if section == "subscription" else None,
            "sort_order": order,
        })
        order += 1

    # ── Loans / installments: detected by a done/ongoing (or boolean) marker
    # cell paired with an amount. This catches both header-backed loan sections
    # and the header-less rows that sit at the bottom of recent sheets.
    lorder = 0
    for row in rows:
        marker_done = None
        for c in row:
            if isinstance(c, bool):
                marker_done = c
                break
            if norm(c) in ("done", "ongoing", "settled"):
                marker_done = norm(c) in ("done", "settled")
                break
        if marker_done is None:
            continue
        name = next((str(c).strip() for c in row if norm(c) and not isinstance(c, bool)
                     and to_num(c) is None and norm(c) not in ("done", "ongoing", "settled")), None)
        amount = next((to_num(c) for c in row if to_num(c) is not None), None)
        if not name or amount is None or amount < 100:
            continue
        if norm(name) in ("card", "utility", "subscription", "merchant") or norm(name).endswith("total"):
            continue
        out["loans"].append({
            "name": name, "amount": amount,
            "status": "done" if marker_done else "ongoing",
            "paid_on": None, "sort_order": lorder,
        })
        lorder += 1

    return out


def import_workbook(path):
    init_db()
    wb = openpyxl.load_workbook(path, data_only=True)
    conn = get_conn()
    try:
        conn.execute("DELETE FROM bills")
        conn.execute("DELETE FROM meters")
        conn.execute("DELETE FROM loans")
        conn.execute("DELETE FROM months")

        summary = []
        for name in wb.sheetnames:
            period = sheet_period(name)
            if not period:
                continue
            data = parse_sheet(grid(wb[name]), period)

            for b in data["bills"]:
                conn.execute(
                    """INSERT INTO bills (period, category, name, amount, min_amount,
                        due_date, status, payment_method, paid_on, paid_amount, cadence, sort_order)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (period, b["category"], b["name"], b["amount"], b["min_amount"],
                     b["due_date"], b["status"], b["payment_method"], b["paid_on"],
                     b["paid_amount"], b["cadence"], b["sort_order"]),
                )
            for m in data["meters"]:
                conn.execute(
                    """INSERT INTO meters (period, name, unit, reading, reading_date, note, sort_order)
                       VALUES (?,?,?,?,?,?,?)""",
                    (period, m["name"], m["unit"], m["reading"], m["reading_date"],
                     m["note"], m["sort_order"]),
                )
            for l in data["loans"]:
                conn.execute(
                    """INSERT INTO loans (period, name, amount, status, paid_on, sort_order)
                       VALUES (?,?,?,?,?,?)""",
                    (period, l["name"], l["amount"], l["status"], l["paid_on"], l["sort_order"]),
                )

            conn.execute("INSERT OR IGNORE INTO months (period) VALUES (?)", (period,))
            summary.append((name, period, len(data["bills"]), len(data["meters"]), len(data["loans"])))
        conn.commit()
    finally:
        conn.close()

    print(f"{'sheet':<22} {'period':<9} bills meters loans")
    print("-" * 50)
    for name, period, nb, nm, nl in summary:
        print(f"{name:<22} {period:<9} {nb:>5} {nm:>6} {nl:>5}")
    print(f"\nImported {len(summary)} months.")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    import_workbook(path)
